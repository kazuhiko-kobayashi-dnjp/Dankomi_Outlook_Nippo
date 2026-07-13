#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tasks.json (data/tasks.json 形式) を、元Excel(◎250217_露出制御業務進捗及び報告.xlsx)を
テンプレートとして書式・条件付き書式・列幅をそのまま維持しつつ、データだけ差し替えて書き出す。

方式: 元ファイル(data/template_source.xlsx)をopenpyxlで読み込み、
      データ行(14行目以降)のセルの値だけを現在のtasksで上書きする。
      ヘッダー・凡例(1〜13行目)・列幅・条件付き書式(担当者別の行色分け、期限の警告色分け、
      工数のカラースケール)・オートフィルタ範囲などはテンプレート側の設定をそのまま利用する。

      なお、テンプレートに含まれるグラフシート(グラフ1/グラフ2)は、openpyxlの
      load_workbook→save→load_workbookの再読込で `Chartsheet.defined_names` エラーが
      発生する既知の不具合があるため、エクスポート版からは除外する
      (元ファイル自体は一切変更しないので、グラフは元ファイル側で引き続き見られる)。

CLIとして: python export_excel.py <data_json_path> <output_xlsx_path>
モジュールとして: from export_excel import build_workbook
"""
import sys
import io
import re
import json
import zipfile
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from richtext_bridge import html_to_richtext

TEMPLATE_PATH = Path(__file__).parent / "data" / "template_source.xlsx"
CHARTSHEETS_TO_REMOVE = ["グラフ2", "グラフ1"]
DATA_START_ROW = 14

# 元Excelの列位置 (Sheet1) との対応
FIELD_COL = {
    "id": 1,        # A: 管理番号
    "project": 2,   # B: 製品・プロジェクト
    "category": 3,  # C: 機能
    "person": 4,    # D: 担当者 (条件付き書式の行色分けの判定列)
    "plan": 5,      # E: 業務の段取り
    "recent": 6,    # F: 直近の動き
    "deadline": 7,  # G: 期限 (条件付き書式の警告色分けの判定列)
    "status_mark": 8,   # H: 打ち上げ
    "manhours": 9,      # I: 工数 (条件付き書式のカラースケール対象列)
    "note2": 10,         # J: 追記メモ(まれに値あり)
    "comment": 13,        # M: コメント
}
HTML_FIELDS = ("plan", "recent", "comment")
# openpyxlが数式と誤認識してしまう先頭文字（"="等で始まる自由記述テキストが
# 実際の業務メモに存在するため、文字列型を明示して破損を防ぐ）
FORMULA_LIKE_PREFIXES = ("=", "+", "-", "@")


def _excel_id_value(id_str):
    """管理番号は元Excelでは数値(int/float)として入っているため、可能な限り数値に戻す。"""
    if id_str in (None, ""):
        return None
    s = str(id_str).strip()
    if not s:
        return None
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return s  # "new171..." のような自動採番IDはそのまま文字列


def _excel_deadline_value(deadline_str):
    """期限はyyyy-mm-dd形式なら日付型に戻す(元の書式=yyyy/m/dで表示される)。
    '完'/'済'等の自由記述はそのまま文字列として保持する(元ファイルも同様の運用)。"""
    if not deadline_str:
        return None
    s = str(deadline_str)
    try:
        return datetime.datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return s


def set_cell_value(cell, val):
    cell.value = val
    if isinstance(val, str) and val.startswith(FORMULA_LIKE_PREFIXES):
        # 先頭が「=」等で始まる自由記述テキストを数式と誤認識させないよう文字列型に固定する
        cell.data_type = "s"


def _cell_value_for_field(key, raw_val):
    if key == "id":
        return _excel_id_value(raw_val)
    if key == "deadline":
        return _excel_deadline_value(raw_val)
    if key in HTML_FIELDS:
        return html_to_richtext(raw_val) if raw_val else None
    if raw_val == "":
        return None
    return raw_val


def _write_task_row(ws, row, task):
    for key, col in FIELD_COL.items():
        val = _cell_value_for_field(key, task.get(key))
        set_cell_value(ws.cell(row, col), val)


def _clear_row(ws, row):
    for col in FIELD_COL.values():
        ws.cell(row, col).value = None


def _clone_row_style(ws, src_row, dst_row, max_col=13):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def build_workbook(tasks):
    """元Excel(template_source.xlsx)を土台に、データ行だけをtasksの内容で差し替えたWorkbookを返す。"""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"テンプレートファイルが見つかりません: {TEMPLATE_PATH}\n"
            "app/data/template_source.xlsx に元Excelのコピーを置いてください。"
        )

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    for name in CHARTSHEETS_TO_REMOVE:
        if name in wb.sheetnames:
            del wb[name]

    ws = wb["Sheet1"]
    template_max_row = ws.max_row  # 編集前に元の最終データ行(462)を記録しておく

    tasks_by_row = {}
    overflow_tasks = []
    for t in tasks:
        row = t.get("row")
        if isinstance(row, int) and DATA_START_ROW <= row <= template_max_row:
            tasks_by_row[row] = t
        else:
            overflow_tasks.append(t)

    # 1) 元テンプレートのデータ行範囲(14〜462行目)を、現在のタスク内容で上書き/クリア
    for row in range(DATA_START_ROW, template_max_row + 1):
        t = tasks_by_row.get(row)
        if t:
            _write_task_row(ws, row, t)
        else:
            _clear_row(ws, row)  # Web上で削除された行は空欄にする(書式・行の高さは維持)

    # 2) Web上で新規追加され、元の行範囲を超えるタスクは末尾に追記
    #    (テンプレートの条件付き書式は826行目まで設定済みのためそのまま適用される)
    overflow_tasks.sort(key=lambda t: t.get("row") or 0)
    next_row = template_max_row + 1
    for t in overflow_tasks:
        _clone_row_style(ws, template_max_row, next_row)
        _write_task_row(ws, next_row, t)
        next_row += 1

    return wb


# ── 破損しやすい静的パーツ(コメント/VML図形/図形描画)の復元 ──────────────────
# openpyxlはSheet1に貼り付けた画像の図形描画(drawing)や、セルコメント用のレガシーVML図形
# (legacyDrawing)を読み込んで保存し直す際、名前空間プレフィックスが崩れたり(v:/o:/x: が
# ns0:/ns1:等に化ける)、拡張情報(a16:creationId等)が失われたりして、Excelで開いたときに
# 「修復が必要」と判定される壊れたXMLを生成してしまうことがある(openpyxl側の既知の制限)。
# これらのパーツはエクスポート処理で値を一切書き換えないため、テンプレート側の元バイト列を
# そのまま上書きコピーすることで破損を回避する。
_STATIC_PART_REL_TYPES = ("/comments", "/vmlDrawing", "/drawing")
_REL_TAG_RE = re.compile(r'<Relationship\b([^>]*)/>')
_REL_ATTR_RE = re.compile(r'(\w+)="([^"]*)"')


def _rel_targets_by_type_suffix(rels_xml: bytes, type_suffix: str) -> list[str]:
    """.rels XMLから、Type が type_suffix で終わるRelationshipのTarget一覧を返す。"""
    text = rels_xml.decode("utf-8")
    targets = []
    for m in _REL_TAG_RE.finditer(text):
        attrs = dict(_REL_ATTR_RE.findall(m.group(1)))
        if attrs.get("Type", "").endswith(type_suffix):
            targets.append(attrs.get("Target", ""))
    return targets


def _resolve_part_path(base_dir: str, target: str) -> str:
    """.rels の Target(相対 or "/"始まりの絶対パス)を zip 内のパーツパスに正規化する。"""
    if target.startswith("/"):
        return target.lstrip("/")
    resolved = []
    for p in (base_dir + "/" + target).split("/"):
        if p == "..":
            if resolved:
                resolved.pop()
        elif p and p != ".":
            resolved.append(p)
    return "/".join(resolved)


def _restore_legacy_drawing_parts(xlsx_bytes: bytes) -> bytes:
    """openpyxlが再シリアライズしたコメント/VML図形/図形描画パーツを、
    テンプレート(template_source.xlsx)の元バイト列で上書きして返す。
    対象パーツが見つからない場合や処理に失敗した場合は、元のバイト列をそのまま返す。"""
    if not TEMPLATE_PATH.exists():
        return xlsx_bytes

    rels_path = "xl/worksheets/_rels/sheet1.xml.rels"
    try:
        with zipfile.ZipFile(TEMPLATE_PATH) as zin:
            names = set(zin.namelist())
            if rels_path not in names:
                return xlsx_bytes
            orig_rels = zin.read(rels_path)
            orig_parts = {}
            for type_suffix in _STATIC_PART_REL_TYPES:
                for target in _rel_targets_by_type_suffix(orig_rels, type_suffix):
                    path = _resolve_part_path("xl/worksheets", target)
                    if path in names:
                        orig_parts[type_suffix] = zin.read(path)

        if not orig_parts:
            return xlsx_bytes

        with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zin2:
            out_names = set(zin2.namelist())
            if rels_path not in out_names:
                return xlsx_bytes
            new_rels = zin2.read(rels_path)
            replace_map = {}
            for type_suffix, orig_bytes in orig_parts.items():
                for target in _rel_targets_by_type_suffix(new_rels, type_suffix):
                    path = _resolve_part_path("xl/worksheets", target)
                    if path in out_names:
                        replace_map[path] = orig_bytes

            if not replace_map:
                return xlsx_bytes

            buf_out = io.BytesIO()
            with zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin2.infolist():
                    data = replace_map.get(item.filename, zin2.read(item.filename))
                    zout.writestr(item, data)
            return buf_out.getvalue()
    except Exception as e:
        print(f"[WARN] レガシー図形パーツの復元をスキップしました: {e}", file=sys.stderr)
        return xlsx_bytes


def export_workbook_bytes(tasks) -> bytes:
    """build_workbook()でWorkbookを作成・保存し、破損しやすい静的パーツ(コメント/VML図形/
    図形描画)をテンプレートの元バイト列で復元したxlsxバイト列を返す。"""
    wb = build_workbook(tasks)
    buf = io.BytesIO()
    wb.save(buf)
    return _restore_legacy_drawing_parts(buf.getvalue())


if __name__ == "__main__":
    DATA_FILE = sys.argv[1] if len(sys.argv) > 1 else "data/tasks.json"
    OUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "export.xlsx"

    with open(DATA_FILE, encoding="utf-8") as f:
        data = json.load(f)

    xlsx_bytes = export_workbook_bytes(data.get("tasks", []))
    with open(OUT_FILE, "wb") as f:
        f.write(xlsx_bytes)
    print(f"Exported {len(data.get('tasks', []))} rows to {OUT_FILE}", file=sys.stderr)


