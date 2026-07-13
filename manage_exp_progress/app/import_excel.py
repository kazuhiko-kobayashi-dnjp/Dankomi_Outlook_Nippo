#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
◎250217_露出制御業務進捗及び報告.xlsx (Sheet1) を data/tasks.json に変換する。

使い方:
    python import_excel.py "<入力xlsxパス>" [出力jsonパス]

シート構造 (Sheet1):
    行1-12  : 凡例・集計欄 (無視)
    行13    : ヘッダー行
    行14以降: データ行
        A: 管理番号
        B: 製品・プロジェクト
        C: 機能
        D: 担当者
        E: 業務の段取り
        F: 直近の動き (ヘッダー文字なし、実質続き列)
        G: 期限
        H: 打ち上げ (ステータス記号: ◎○〇●済完)
        I: 工数
        J: 追記メモ (まれに値あり)
        M: コメント
"""
import sys
import json
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from richtext_bridge import celltext_to_html

HEADER_ROW = 13
DATA_START_ROW = 14

STATUS_LABELS = {
    "◎": "重要",
    "○": "報告あり",
    "〇": "報告あり",
    "●": "順調",
    "済": "完了",
    "完": "完了",
    "▲": "遅延しそう",
}


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def cell_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s[:10] if s else None


def cell_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_tasks_from_excel(in_path):
    """Excel(Sheet1)を読み込み、(tasksリスト, meta dict)を返す。"""
    wb = openpyxl.load_workbook(in_path, data_only=True, rich_text=True)
    ws = wb["Sheet1"]

    tasks = []
    projects = set()
    categories = set()
    persons = set()

    for row in range(DATA_START_ROW, ws.max_row + 1):
        vals = [ws.cell(row, c).value for c in range(1, 14)]
        task_id, project, category, person, plan, recent, deadline, status_mark, manhours = vals[0:9]
        note2 = vals[9]
        comment = vals[12]
        plan_font = ws.cell(row, 5).font
        recent_font = ws.cell(row, 6).font
        comment_font = ws.cell(row, 13).font

        # 完全に空行はスキップ
        if not any(v not in (None, "") for v in vals):
            continue

        project_s = cell_str(project)
        category_s = cell_str(category)
        person_s = cell_str(person)
        status_mark_s = cell_str(status_mark)

        if project_s:
            projects.add(project_s)
        if category_s:
            categories.add(category_s)
        if person_s:
            persons.add(person_s)

        tasks.append({
            "id": cell_str(task_id) or f"row{row}",
            "row": row,
            "project": project_s,
            "category": category_s,
            "person": person_s,
            "plan": celltext_to_html(plan, plan_font),
            "recent": celltext_to_html(recent, recent_font),
            "deadline": cell_date(deadline),
            "status_mark": status_mark_s,
            "status": STATUS_LABELS.get(status_mark_s, "" if status_mark_s else "未設定"),
            "manhours": cell_num(manhours),
            "note2": cell_str(note2),
            "comment": celltext_to_html(comment, comment_font),
            "created_at": None,
            "updated_at": None,
        })

    meta = {"projects": sorted(projects), "categories": sorted(categories), "persons": sorted(persons)}
    return tasks, meta


# 差分検出対象のフィールド(created_at/updated_at/rowは除く)
_COMPARE_FIELDS = ("id", "project", "category", "person", "plan", "recent",
                   "deadline", "status_mark", "status", "manhours", "note2", "comment")


def refresh_tasks_from_excel(source_path, existing_data):
    """元Excelを再読み込みし、既存のtasks.json(existing_data)とマージする。

    - Excel側で内容が変わっていればその行を更新(updated_atを現在時刻に)
    - Excelにのみ存在する新しい行は追加
    - Web側でのみ作成されたタスク(Excelの行範囲外)はそのまま保持
    - 戻り値: (新しいdata dict, サマリ dict{added, updated, unchanged})
    """
    fresh_tasks, _ = read_tasks_from_excel(source_path)
    fresh_by_row = {t["row"]: t for t in fresh_tasks}
    existing_by_row = {t["row"]: t for t in existing_data.get("tasks", [])}

    now = datetime.datetime.now().isoformat(timespec="seconds")
    added = updated = unchanged = 0
    result_tasks = []

    for row in sorted(fresh_by_row):
        fresh = fresh_by_row[row]
        old = existing_by_row.get(row)
        if old is None:
            fresh["created_at"] = now
            fresh["updated_at"] = now
            result_tasks.append(fresh)
            added += 1
        else:
            changed = any(old.get(k) != fresh.get(k) for k in _COMPARE_FIELDS)
            if changed:
                merged = {**old, **fresh}
                merged["created_at"] = old.get("created_at") or now
                merged["updated_at"] = now
                result_tasks.append(merged)
                updated += 1
            else:
                result_tasks.append(old)
                unchanged += 1

    # Excelの行範囲外(Web上で新規作成したタスクなど)はそのまま測す
    for row, old in existing_by_row.items():
        if row not in fresh_by_row:
            result_tasks.append(old)

    result_tasks.sort(key=lambda t: t["row"])

    new_data = dict(existing_data)
    new_data["tasks"] = result_tasks
    summary = {"added": added, "updated": updated, "unchanged": unchanged}
    return new_data, summary


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <input.xlsx> [output.json]", file=sys.stderr)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent / "data" / "tasks.json"

    tasks, meta = read_tasks_from_excel(in_path)

    out = {
        "meta": {
            "source_file": str(in_path.name),
            "imported_at": datetime.datetime.now().isoformat(timespec="seconds"),
            **meta,
        },
        "tasks": tasks,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Imported {len(tasks)} tasks -> {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
