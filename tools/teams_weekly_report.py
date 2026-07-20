# -*- coding: utf-8 -*-
"""
Teams チャット週報自動生成スクリプト

分析対象:
  - 自分が発信したチャット/チャンネルメッセージ
  - 自分がメンションされたメッセージ

使い方:
  py tools/teams_weekly_report.py                  # 直近7日 / 対話認証
  py tools/teams_weekly_report.py --days 14        # 直近14日
  py tools/teams_weekly_report.py --device-code    # Device Code Flow
  py tools/teams_weekly_report.py --no-channels    # チャンネル除外（高速）
  py tools/teams_weekly_report.py --out report.md  # 出力先指定

必要パッケージ:
  pip install msal requests beautifulsoup4 lxml
"""

import sys
import os
import json
import re
import time
import argparse
import subprocess
from datetime import datetime, timezone, timedelta
from pathlib import Path
from html.parser import HTMLParser

# ── プロキシ自動検出 ─────────────────────────────────────────────────
def _auto_detect_proxy():
    if os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY"):
        return
    try:
        ps_cmd = (
            '$p=[System.Net.WebRequest]::GetSystemWebProxy();'
            '$u=[Uri]"https://login.microsoftonline.com";'
            '$r=$p.GetProxy($u);'
            'if($r.Authority -ne $u.Authority){Write-Output $r.AbsoluteUri}'
        )
        r = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=10,
        )
        proxy = r.stdout.strip().rstrip("/")
        if proxy:
            os.environ["HTTP_PROXY"] = proxy
            os.environ["HTTPS_PROXY"] = proxy
            print(f"[INFO] Proxy自動検出: {proxy}", flush=True)
            return
    except Exception:
        pass
    try:
        r = subprocess.run(["netsh", "winhttp", "show", "proxy"],
                           capture_output=True, text=True, timeout=5)
        for line in r.stdout.splitlines():
            if "Proxy Server" in line and ":" in line:
                proxy = line.split(":", 1)[1].strip()
                if proxy and proxy != "(null)":
                    os.environ["HTTP_PROXY"] = proxy
                    os.environ["HTTPS_PROXY"] = proxy
                    print(f"[INFO] Proxy(WinHTTP): {proxy}", flush=True)
                    return
    except Exception:
        pass

_auto_detect_proxy()

import msal
import requests

# ── LLM APIキー管理 ──────────────────────────────────────────────────
_KEY_FILE = Path(__file__).with_name(".gh_models_token")

def _load_api_key(explicit: str = "") -> str:
    """優先順位: --api-key引数 > GH_MODELS_TOKEN環境変数 > .gh_models_tokenファイル > OPENAI_API_KEY環境変数"""
    if explicit.strip():
        return explicit.strip()
    v = os.environ.get("GH_MODELS_TOKEN", "")
    if v:
        return v
    if _KEY_FILE.exists():
        return _KEY_FILE.read_text(encoding="utf-8").strip()
    return os.environ.get("OPENAI_API_KEY", "")

def _save_api_key(key: str):
    """トークンを .gh_models_token ファイルに保存"""
    _KEY_FILE.write_text(key.strip(), encoding="utf-8")
    print(f"[INFO] APIキーを保存しました: {_KEY_FILE}", flush=True)

# ── ユーザー固有設定(担当者名・業務進捗Excelパス等) ────────────────────
# git管理対象外(.gitignore)のファイル。このファイルがなければ何も自動設定されず、
# コマンド引数(--task-owner 等)を明示しない限り業務進捗表機能は無効のままとなる。
# 設定例: tools/.user_config.example.json を参照。
_USER_CONFIG_FILE = Path(__file__).with_name(".user_config.json")

def _load_user_config() -> dict:
    if not _USER_CONFIG_FILE.exists():
        return {}
    try:
        return json.loads(_USER_CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[WARN] {_USER_CONFIG_FILE} の読み込みに失敗しました: {e}", flush=True)
        return {}

def _save_user_config(**updates):
    """既存設定に updates をマージして .user_config.json に保存"""
    cfg = _load_user_config()
    cfg.update({k: v for k, v in updates.items() if v})
    _USER_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[INFO] 個人設定を保存しました: {_USER_CONFIG_FILE}", flush=True)

# ── 収集済みメッセージのキャッシュ（Graph API再取得なしで -Summarize を後から実行するため） ──
_RECORDS_CACHE_FILE = Path(__file__).parent.parent / "output" / ".records_cache.json"

def _save_records_cache(records: list[dict], days: int, me_name: str, me_id: str, since: datetime):
    """収集したrecords一式をJSONでキャッシュ保存する（dtはISO文字列化）。"""
    try:
        _RECORDS_CACHE_FILE.parent.mkdir(exist_ok=True)
        payload = {
            "days": days,
            "me_name": me_name,
            "me_id": me_id,
            "since": since.isoformat(),
            "saved_at": datetime.now(JST).isoformat(),
            "records": [
                {**{k: v for k, v in r.items() if k != "dt"}, "dt": r["dt"].isoformat()}
                for r in records
            ],
        }
        _RECORDS_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"[WARN] 収集結果のキャッシュ保存に失敗しました: {e}", flush=True)

def _load_records_cache(path: Path = _RECORDS_CACHE_FILE) -> dict:
    """キャッシュされたrecords一式を読み込む（dtはdatetimeに復元）。無ければ例外送出。"""
    data = json.loads(path.read_text(encoding="utf-8"))
    for r in data["records"]:
        r["dt"] = datetime.fromisoformat(r["dt"])
    return data

# ── 設定 ──────────────────────────────────────────────────────────────
AUTHORITY    = "https://login.microsoftonline.com/69405920-b673-4f7c-8845-e124e9d08af2"
CLIENT_ID    = "d3590ed6-52b3-4102-aeff-aad2292ab01c"   # Microsoft Office (1st party)
SCOPES       = ["https://graph.microsoft.com/.default"]
GRAPH        = "https://graph.microsoft.com/v1.0"
_CACHE_FILE  = Path(__file__).with_name(".msal_token_cache.json")

JST = timezone(timedelta(hours=9))

# ── 認証 ──────────────────────────────────────────────────────────────
def _load_cache():
    cache = msal.SerializableTokenCache()
    if _CACHE_FILE.exists():
        cache.deserialize(_CACHE_FILE.read_text(encoding="utf-8"))
    return cache

def _save_cache(cache):
    if cache.has_state_changed:
        _CACHE_FILE.write_text(cache.serialize(), encoding="utf-8")

def get_token(auth_mode: str) -> str:
    print("[INFO] MSAL 初期化中...", flush=True)
    cache = _load_cache()
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)

    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            print("  キャッシュトークン使用", flush=True)
            _save_cache(cache)
            return result["access_token"]

    if auth_mode == "device_code":
        flow = app.initiate_device_flow(scopes=SCOPES)
        print("=" * 55, flush=True)
        print(f"  認証URL : {flow['verification_uri']}", flush=True)
        print(f"  コード  : {flow['user_code']}", flush=True)
        print("=" * 55, flush=True)
        print("ブラウザで上記URLを開きコードを入力してください...", flush=True)
        result = app.acquire_token_by_device_flow(flow)
    else:
        result = app.acquire_token_interactive(scopes=SCOPES)

    if "access_token" not in result:
        print(f"[ERROR] 認証失敗: {result.get('error_description', '')[:200]}", flush=True)
        sys.exit(1)

    print("  認証成功", flush=True)
    _save_cache(cache)
    return result["access_token"]

# ── Graph クライアント ────────────────────────────────────────────────
class GraphClient:
    def __init__(self, token: str):
        self.s = requests.Session()
        self.s.headers.update({
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        })
        self._last_status: int = 0        # 直近リクエストの HTTP ステータス
        self._first_page_status: int = 0  # get_paged の1ページ目ステータス

    def get(self, path: str, params: dict = None, silent: bool = False):
        url = path if path.startswith("http") else f"{GRAPH}{path}"
        resp = None
        for attempt in range(3):
            resp = self.s.get(url, params=params)
            self._last_status = resp.status_code
            if resp.status_code == 200:
                return resp.json()
            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", "10"))
                print(f"  [WAIT] 429 レート制限, {wait}秒待機...", flush=True)
                time.sleep(wait)
                continue
            break
        if not silent:
            try:
                err = resp.json().get("error", {})
                print(f"  [WARN] {resp.status_code} {err.get('code','')}: {err.get('message','')[:120]}", flush=True)
            except Exception:
                print(f"  [WARN] {resp.status_code}: {path.split('?')[0][:80]}", flush=True)
        return None

    def get_paged(self, path: str, params: dict = None, limit: int = 500):
        """@odata.nextLink を辿って全ページ取得"""
        items = []
        url = path if path.startswith("http") else f"{GRAPH}{path}"
        self._first_page_status = 0
        first = True
        while url and len(items) < limit:
            data = self.get(url, params=params)
            if first:
                self._first_page_status = self._last_status
                first = False
            params = None  # 2ページ目以降はnextLinkにパラメータが含まれる
            if not data:
                break
            items.extend(data.get("value", []))
            url = data.get("@odata.nextLink")
            if url:
                time.sleep(0.3)
        return items

# ── HTML → プレーンテキスト ──────────────────────────────────────────
class _HtmlStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.texts = []
        self._skip = False
    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style"):
            self._skip = True
        if tag == "br":
            self.texts.append("\n")
        if tag == "p":
            self.texts.append("\n")
    def handle_endtag(self, tag):
        if tag in ("script", "style"):
            self._skip = False
    def handle_data(self, data):
        if not self._skip:
            self.texts.append(data)

def html_to_text(html: str) -> str:
    if not html:
        return ""
    p = _HtmlStripper()
    p.feed(html)
    text = "".join(p.texts)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

# ── メッセージ解析ユーティリティ ─────────────────────────────────────
def parse_dt(s: str) -> datetime:
    """ISO 8601 文字列 → JST datetime"""
    if not s:
        return datetime.min.replace(tzinfo=JST)
    s = re.sub(r"(\.\d{3})\d+Z$", r"\1Z", s)  # マイクロ秒を丸める
    dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
    return dt.astimezone(JST)

def extract_mentions(msg: dict) -> list[str]:
    """mentions配列からdisplayNameのリストを返す"""
    result = []
    for m in (msg.get("mentions") or []):
        if not isinstance(m, dict):
            continue
        mentioned = m.get("mentioned") or {}
        user = mentioned.get("user") or {}
        name = user.get("displayName", "")
        if name:
            result.append(name)
    return result

def msg_is_sent_by(msg: dict, user_id: str) -> bool:
    from_data = msg.get("from") or {}
    user_data = from_data.get("user") or {}
    return user_data.get("id") == user_id

def msg_mentions(msg: dict, user_id: str) -> bool:
    for m in (msg.get("mentions") or []):
        if not isinstance(m, dict):
            continue
        mentioned = m.get("mentioned") or {}
        user = mentioned.get("user") or {}
        if user.get("id") == user_id:
            return True
    return False

def msg_summary(msg: dict) -> str:
    """メッセージ本文を短いテキストに変換"""
    body = msg.get("body", {})
    if body.get("contentType") == "html":
        text = html_to_text(body.get("content", ""))
    else:
        text = body.get("content", "")
    # システムメッセージ除去
    if msg.get("messageType") != "message":
        return ""
    # 長すぎる場合は切り詰め
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    text = " / ".join(lines[:4])
    if len(text) > 200:
        text = text[:197] + "..."
    return text

# ── データ収集 ───────────────────────────────────────────────────────
def collect_chat_messages(gc: GraphClient, me_id: str, since: datetime) -> list[dict]:
    """
    1:1 / グループチャットのメッセージを収集。
    対象: 自分が発信 OR 自分がメンションされた。
    """
    print("[INFO] チャット一覧を取得中...", flush=True)
    since_str = since.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    # NOTE: $filter=lastUpdatedDateTime は oneOnOne チャットを返さない Graph API の挙動があるため
    # サーバー側フィルタを使わず最新200件を取得してクライアント側で絞る
    chats_raw = gc.get_paged(
        "/me/chats",
        params={"$select": "id,chatType,topic,lastUpdatedDateTime",
                "$top": "50"},
        limit=200,
    )
    # chatType別の取得分布を診断表示
    type_dist: dict[str, int] = {}
    for c in chats_raw:
        ct = c.get("chatType", "unknown")
        type_dist[ct] = type_dist.get(ct, 0) + 1
    print(f"  チャット数: {len(chats_raw)} 件取得 (内訳: {type_dist})", flush=True)

    # NOTE: oneOnOne / group チャットは lastUpdatedDateTime が Graph API 側でバグにより
    # 実際のメッセージ日時を反映しない（更新されない）ことがある（既知の挙動。実測で
    # groupチャット37件中31件がこのフィルタで誤って除外されるケースを確認済み・2026-07-13）。
    # そのため oneOnOne / group は全件取得してメッセージ日付でフィルタする。
    # meeting は件数が非常に多い（100件超）ため、性能維持のため lastUpdatedDateTime での
    # 事前フィルタを維持する（誤除外のリスクは残るが許容）。
    chats = []
    for c in chats_raw:
        ct = c.get("chatType", "")
        if ct in ("oneOnOne", "group"):
            chats.append(c)  # メタデータに関わらず全件チェック
        elif (c.get("lastUpdatedDateTime") or "") >= since_str:
            chats.append(c)

    # 期間内のchatType別内訳
    type_in: dict[str, int] = {}
    for c in chats:
        ct = c.get("chatType", "unknown")
        type_in[ct] = type_in.get(ct, 0) + 1
    print(f"  対象: {len(chats)} 件 (内訳: {type_in})", flush=True)

    records = []
    # chatType別の成否を記録: {ctype: {"ok": [topic,...], "err": [(status,topic),...]}}
    type_stats: dict = {}

    for i, chat in enumerate(chats, 1):
        cid   = chat["id"]
        ctype = chat.get("chatType", "unknown")
        topic = chat.get("topic") or (
            "1:1チャット" if ctype == "oneOnOne" else "グループチャット"
        )
        if ctype not in type_stats:
            type_stats[ctype] = {"ok": [], "err": []}
        ctype_label = {"oneOnOne": "1on1", "group": "group", "meeting": "mtg"}.get(ctype, ctype[:4])
        print(f"  [{i}/{len(chats)}][{ctype_label}] {topic[:30]}", end=" ", flush=True)

        # NOTE: /me/chats/{id}/messages は $filter/$orderby 非サポート
        # $select も 500 を誘発するため除去し、クライアント側でフィルタ
        msgs = gc.get_paged(
            f"/me/chats/{cid}/messages",
            params={"$top": "50"},
            limit=300,
        )
        st = gc._first_page_status
        if st not in (200, 0):
            type_stats[ctype]["err"].append((st, topic))
        else:
            type_stats[ctype]["ok"].append(topic)

        # 1:1チャットは自分が参加しているなら相手の発言も取得（メンション不要）
        # グループ/会議チャットは sent/mentioned のみ（ノイズ抑制）
        is_1on1 = (ctype == "oneOnOne")

        # 1:1の場合: 期間内に自分が発言しているか事前確認（参加済み会話のみ）
        user_participated = any(
            msg_is_sent_by(m, me_id)
            for m in msgs
            if isinstance(m, dict) and parse_dt(m.get("createdDateTime", "")) >= since
        ) if is_1on1 else False

        hit = 0
        for msg in msgs:
            if not isinstance(msg, dict):
                continue
            dt = parse_dt(msg.get("createdDateTime", ""))
            if dt < since:
                continue
            is_sent = msg_is_sent_by(msg, me_id)
            is_mentioned = msg_mentions(msg, me_id)

            # 抽出条件:
            #   1:1 かつ自分が発言済み → 相手の発言も含める（返信・成果報告を取りこぼさない）
            #   グループ/会議 → sent/mentioned のみ
            if not (is_1on1 and user_participated):
                if not (is_sent or is_mentioned):
                    continue

            text = msg_summary(msg)
            if not text:
                continue

            if is_sent:
                subtype = "sent"
            elif is_mentioned:
                subtype = "mentioned"
            else:
                subtype = "other"  # 1:1での相手発言

            records.append({
                "type":      "chat",
                "subtype":   subtype,
                "source":    topic,
                "dt":        dt,
                "date":      dt.strftime("%Y-%m-%d"),
                "time":      dt.strftime("%H:%M"),
                "text":      text,
            })
            hit += 1
        print(f"({hit}件)", flush=True)
        time.sleep(0.2)

    # ── chatType 別 層別レポート ──────────────────────────────────────
    print("\n[INFO] チャット取得結果 (chatType 層別):", flush=True)
    for ctype, s in sorted(type_stats.items()):
        ok_n, err_n = len(s["ok"]), len(s["err"])
        print(f"  {ctype:10s}: 成功 {ok_n:3d} / 失敗 {err_n:3d}", flush=True)
        if err_n:
            # 失敗チャットを最大5件表示
            for status, t in s["err"][:5]:
                print(f"    [{status}] {t[:45]}", flush=True)
            if err_n > 5:
                print(f"    ...他 {err_n-5} 件", flush=True)

    return records


_TEAMS_ALLOW_FILE = Path(__file__).with_name(".teams_allow")

def list_teams(gc: GraphClient):
    """参加中の全 Teams とチャンネルを表示して終了"""
    print("[INFO] 参加 Teams を取得中...", flush=True)
    teams = gc.get_paged("/me/joinedTeams",
                         params={"$select": "id,displayName"}, limit=50)
    print(f"\n参加 Teams 一覧 ({len(teams)} 件):\n", flush=True)
    for team in sorted(teams, key=lambda t: t.get("displayName", "")):
        tname = team.get("displayName", "")
        channels = gc.get_paged(
            f"/teams/{team['id']}/channels",
            params={"$select": "displayName"}, limit=50,
        )
        chnames = ", ".join(c.get("displayName", "") for c in channels)
        print(f"  [{tname}]  channels: {chnames}", flush=True)
    print("\n--teams-allow に指定したいチーム名（部分一致・カンマ区切り）をコピーしてください。", flush=True)
    print("例: .\\wr.ps1 -Channels -TeamsAllow \"BEV,FCM,DCAP\"", flush=True)


def _load_teams_allow(explicit: str = "") -> list[str]:
    """優先順位: --teams-allow引数 > .teams_allow ファイル"""
    raw = explicit.strip()
    if not raw and _TEAMS_ALLOW_FILE.exists():
        raw = _TEAMS_ALLOW_FILE.read_text(encoding="utf-8").strip()
    if not raw:
        return []
    return [s.strip() for s in raw.split(",") if s.strip()]


def collect_channel_messages(
    gc: GraphClient, me_id: str, since: datetime,
    teams_allow: list[str] | None = None,
) -> list[dict]:
    """
    参加 Teams チャンネルのメッセージを収集。
    teams_allow: チーム名の部分一致フィルタ（空リストなら全Teams対象）
    """
    print("[INFO] 参加 Teams を取得中...", flush=True)
    teams = gc.get_paged("/me/joinedTeams",
                         params={"$select": "id,displayName"}, limit=50)
    print(f"  Teams 数: {len(teams)}", flush=True)

    # フィルタ適用
    if teams_allow:
        allow_lower = [a.lower() for a in teams_allow]
        teams = [t for t in teams
                 if any(a in t.get("displayName", "").lower() for a in allow_lower)]
        print(f"  フィルタ後: {len(teams)} 件 ({', '.join(teams_allow)})", flush=True)

    records = []

    for team in teams:
        tid   = team["id"]
        tname = team.get("displayName", "Unknown Team")

        # lastMessagePreview を取得して非活性チャンネルを事前除外
        channels_raw = gc.get_paged(
            f"/teams/{tid}/channels",
            params={"$select": "id,displayName", "$expand": "lastMessagePreview"}, limit=50,
        )
        since_str = since.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        channels = []
        skipped_ch = 0
        for ch in channels_raw:
            lmp = ch.get("lastMessagePreview") or {}
            lmp_dt = lmp.get("createdDateTime", "")
            if lmp_dt and lmp_dt < since_str:
                skipped_ch += 1
            else:
                channels.append(ch)
        if skipped_ch:
            print(f"  [{tname[:25]}] 非活性チャンネル {skipped_ch} 件スキップ / 残り {len(channels)} 件", flush=True)

        for ch in channels:
            chid   = ch["id"]
            chname = ch.get("displayName", "")
            source = f"{tname} > {chname}"
            print(f"  チャンネル: {source[:50]}", end=" ", flush=True)

            # NOTE: チャンネルメッセージも $select が 500 を誘発するため除去
            msgs = gc.get_paged(
                f"/teams/{tid}/channels/{chid}/messages",
                params={"$top": "30"},
                limit=150,
            )
            hit = 0
            for msg in msgs:
                if not isinstance(msg, dict):
                    continue
                dt = parse_dt(msg.get("createdDateTime", ""))
                if dt < since:
                    continue
                for m in [msg]:  # 本文
                    is_sent = msg_is_sent_by(m, me_id)
                    is_mentioned = msg_mentions(m, me_id)
                    if is_sent or is_mentioned:
                        text = msg_summary(m)
                        if text:
                            records.append({
                                "type":    "channel",
                                "subtype": "sent" if is_sent else "mentioned",
                                "source":  source,
                                "dt":      dt,
                                "date":    dt.strftime("%Y-%m-%d"),
                                "time":    dt.strftime("%H:%M"),
                                "text":    text,
                            })
                            hit += 1

                # リプライ確認（親メッセージが自分に関連する場合のみ）
                if msg_is_sent_by(msg, me_id) or msg_mentions(msg, me_id):
                    replies = gc.get_paged(
                        f"/teams/{tid}/channels/{chid}/messages/{msg['id']}/replies",
                        params={"$top": "20"},
                        limit=50,
                    )
                else:
                    replies = []
                for reply in replies:
                    rdt = parse_dt(reply.get("createdDateTime", ""))
                    if rdt < since:
                        continue
                    is_sent = msg_is_sent_by(reply, me_id)
                    is_mentioned = msg_mentions(reply, me_id)
                    if not (is_sent or is_mentioned):
                        continue
                    text = msg_summary(reply)
                    if not text:
                        continue
                    records.append({
                        "type":    "channel-reply",
                        "subtype": "sent" if is_sent else "mentioned",
                        "source":  source,
                        "dt":      rdt,
                        "date":    rdt.strftime("%Y-%m-%d"),
                        "time":    rdt.strftime("%H:%M"),
                        "text":    f"[reply] {text}",
                    })
                    hit += 1

            print(f"({hit}件)", flush=True)
            time.sleep(0.3)

    return records

# ── 週報レポート生成 ──────────────────────────────────────────────────
def build_report(records: list[dict], me_name: str, days: int,
                 ai_summary: str = "", nippo_text: str = "") -> str:
    if not records:
        if nippo_text.strip():
            return (
                "（対象期間中に該当メッセージが見つかりませんでした）\n\n"
                "## 日報（業務実績抜粋）\n\n" + nippo_text + "\n"
            )
        return "（対象期間中に該当メッセージが見つかりませんでした）\n"

    # 日付順にソート
    records.sort(key=lambda r: r["dt"])

    today     = datetime.now(JST)
    week_end  = today.strftime("%Y/%m/%d")
    week_start = (today - timedelta(days=days)).strftime("%Y/%m/%d")
    sent_n    = len([r for r in records if r["subtype"] == "sent"])
    mntnd_n   = len([r for r in records if r["subtype"] == "mentioned"])

    lines = []
    lines.append(f"# 週報 候補案　{week_start} ～ {week_end}")
    lines.append(f"> 分析対象: {me_name} が発信 or メンションされたメッセージ")
    lines.append(f"> 統計: 総件数 {len(records)} 件（発信 {sent_n} / 被メンション {mntnd_n}）  生成: {today.strftime('%Y-%m-%d %H:%M')} JST")
    lines.append("")

    # ─ 日報（業務実績抜粋） ─
    if nippo_text.strip():
        lines.append("## 日報（業務実績抜粋）")
        lines.append("")
        lines.append(nippo_text)
        lines.append("")
        lines.append("---")
        lines.append("")

    # AI 要約がある場合はそれのみを出力（ログ羅列はスキップ）
    if ai_summary:
        lines.append("## AI 要約")
        lines.append("")
        lines.append(ai_summary)
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## 来週の予定・メモ")
        lines.append("")
        lines.append("- （記入）")
        lines.append("")
        return "\n".join(lines) + "\n"

    # AI 要約なし: 日別ログ + テンプレートを出力
    # ─ 1. 日付別サマリー ─
    lines.append("## 1. 日別 活動ログ")
    lines.append("")

    from itertools import groupby
    for date, grp in groupby(records, key=lambda r: r["date"]):
        items = list(grp)
        dt_obj = datetime.strptime(date, "%Y-%m-%d")
        weekdays = "月火水木金土日"
        wd = weekdays[dt_obj.weekday()]
        lines.append(f"### {date}（{wd}）")
        for r in items:
            icon = "→" if r["subtype"] == "sent" else "←"
            label = "発信" if r["subtype"] == "sent" else "被メンション"
            lines.append(f"- `{r['time']}` [{label}] **{r['source'][:40]}**")
            lines.append(f"  {icon} {r['text']}")
        lines.append("")

    # ─ 2. チャネル/チャット別サマリー ─
    lines.append("## 2. チャット/チャンネル別 サマリー")
    lines.append("")

    source_groups: dict[str, list] = {}
    for r in records:
        source_groups.setdefault(r["source"], []).append(r)

    for src, items in sorted(source_groups.items(),
                              key=lambda kv: -len(kv[1])):
        sent  = [r for r in items if r["subtype"] == "sent"]
        mntnd = [r for r in items if r["subtype"] == "mentioned"]
        lines.append(f"### {src}")
        lines.append(f"- 発信: {len(sent)}件 / 被メンション: {len(mntnd)}件")
        # 最近3件だけ掲載
        for r in items[-3:]:
            icon = "→" if r["subtype"] == "sent" else "←"
            lines.append(f"  - `{r['date']} {r['time']}` {icon} {r['text']}")
        lines.append("")

    # ─ 3. 週報本文 テンプレート ─
    lines.append("## 3. 週報 テンプレート（要編集）")
    lines.append("")
    lines.append("### 今週の主な活動")
    lines.append("")

    # ソース別に1行ずつ箇条書きの骨格を生成
    for src, items in sorted(source_groups.items(),
                              key=lambda kv: -len(kv[1])):
        sent_count = len([r for r in items if r["subtype"] == "sent"])
        if sent_count == 0:
            continue
        last = [r for r in items if r["subtype"] == "sent"][-1]
        lines.append(f"- 【{src}】{last['text']}")
    lines.append("")
    lines.append("### 来週の予定")
    lines.append("")
    lines.append("- （記入）")
    lines.append("")
    lines.append("### 課題・懸念事項")
    lines.append("")
    lines.append("- （記入）")
    lines.append("")

    # ─ 4. 統計 ─
    lines.append("---")
    lines.append(f"**統計**: 総件数 {len(records)} 件  "
                 f"（発信 {len([r for r in records if r['subtype']=='sent'])} / "
                 f"被メンション {len([r for r in records if r['subtype']=='mentioned'])}）")
    lines.append(f"生成日時: {today.strftime('%Y-%m-%d %H:%M')} JST")

    return "\n".join(lines) + "\n"

# ── 日報読み込み ──────────────────────────────────────────────────────
def load_nippo_text(nippo_dir: Path, since: datetime, until: datetime) -> str:
    """日報ファイルから＜業務実績＞セクションを抽出してテキストで返す"""
    results = []
    dt = since.replace(hour=0, minute=0, second=0, microsecond=0)
    end = until.replace(hour=23, minute=59, second=59)
    while dt <= end:
        fname = dt.strftime("%y%m%d") + "_日報.txt"
        fpath = nippo_dir / fname
        if fpath.exists():
            for enc in ("utf-8-sig", "cp932", "utf-8"):
                try:
                    text = fpath.read_text(encoding=enc, errors="replace")
                    break
                except Exception:
                    continue
            else:
                dt += timedelta(days=1)
                continue

            if "＜業務実績＞" in text:
                section = text.split("＜業務実績＞", 1)[1]
                meaningful = []
                for line in section.splitlines():
                    s = line.strip()
                    if not s or s in ("・", "＜＞", "＊"):
                        continue
                    # ファイルパス・URL・環境変数・PAT/APIキー行は除外
                    if re.match(r"^(\$env:|file://|https?://|http://|\\\\|[a-z]:\\)", s, re.I):
                        continue
                    # PAT/APIキーらしき行を除外（ghp_xxx、UUIDなど）
                    if re.match(r"^(ghp_|gho_|github_pat_|[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4})", s, re.I):
                        continue
                    # 時刻マーカー（"06:30　＜...＞"）は見出しとして保持
                    if re.match(r"^\d{2}:\d{2}", s):
                        meaningful.append(s)
                    else:
                        meaningful.append(s)

                content = "\n".join(meaningful).strip()
                if content:
                    results.append(f"=== {dt.strftime('%Y-%m-%d')} ===")
                    results.append(content[:2000])
        dt += timedelta(days=1)

    return "\n\n".join(results)


# ── Excel業務進捗表読み込み ──────────────────────────────────────────
def _xl_date(val) -> str:
    """ExcelのシリアルDate/datetimeオブジェクト/int を YYYY/MM/DD 文字列に変換"""
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, (int, float)) and val > 1:
        try:
            from datetime import date, timedelta as td
            d = date(1899, 12, 30) + td(days=int(val))
            return d.strftime("%Y/%m/%d")
        except Exception:
            pass
    return str(val) if val else ""


def load_task_excel(xlsx_path: str, name_filter: str = "") -> list[dict]:
    """業務進捗Excelから担当者=name_filterの行を読み込む（Sheet1を使用）。
    name_filterが空の場合は担当者で絞り込まず全行を対象とする。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"[WARN] Excel読み込み失敗: {e}", flush=True)
        return []

    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # ヘッダー行を探す（'担当者'セルが存在する行）
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_col=2, max_col=9, values_only=True), 1):
        if any("担当者" in str(v) for v in row if v):
            header_row = ri
            break
    if header_row is None:
        print("[WARN] Excel: ヘッダー行が見つかりません", flush=True)
        return []

    tasks = []
    for row in ws.iter_rows(min_row=header_row + 1, min_col=2, max_col=9, values_only=True):
        if not any(v for v in row):
            continue
        b, c, d, e, f, g, h, iv = (row[k] if k < len(row) else None for k in range(8))
        if not d or name_filter not in str(d):
            continue
        if not e and not f:
            continue
        tasks.append({
            "製品":      str(b or ""),
            "業務ジャンル": str(c or ""),
            "担当者":    str(d or ""),
            "段取り":    str(e or "")[:200],
            "進捗":      str(f or "")[:400],
            "期限":      _xl_date(g),
            "打ち上げ":  str(h or ""),
            "工数(H)":   str(iv) if iv else "",
        })
    owner_label = name_filter or "(全員)"
    print(f"[INFO] 業務進捗表: {len(tasks)} 行 (担当者={owner_label})", flush=True)
    return tasks


# ── manage_exp_progress (Web化システム) の tasks.json 読み込み ──────────────
_TASK_HTML_TAG_BLOCK_RE = re.compile(r"</?(?:div|p|br)[^>]*>", re.IGNORECASE)
_TASK_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _html_to_plain(text: str) -> str:
    """manage_exp_progressのリッチテキスト(HTML)欄をプレーンテキストに変換する。
    画像(<img>タグ、貼り付け機能で挿入されたdata URL含む)はタグごと除去されるため、
    本文には影響しない。"""
    if not text:
        return ""
    if "<" not in text:
        return text
    import html as _html
    t = _TASK_HTML_TAG_BLOCK_RE.sub("\n", text)
    t = _TASK_HTML_TAG_RE.sub("", t)
    return _html.unescape(t)


def load_task_json(json_path: str, name_filter: str = "") -> list[dict]:
    """manage_exp_progress(Web化システム)のdata/tasks.jsonから担当者=name_filterの行を読み込む。
    load_task_excel()と同じ戻り値の形(キー名)にする。
    name_filterが空の場合は担当者で絞り込まず全行を対象とする。"""
    import json as _json
    try:
        with open(json_path, encoding="utf-8") as f:
            data = _json.load(f)
    except Exception as e:
        print(f"[WARN] tasks.json読み込み失敗: {e}", flush=True)
        return []

    tasks = []
    for t in data.get("tasks", []):
        person = t.get("person") or ""
        if name_filter and name_filter not in person:
            continue
        if not t.get("plan") and not t.get("recent"):
            continue
        manhours = t.get("manhours")
        tasks.append({
            "製品":      t.get("project") or "",
            "業務ジャンル": t.get("category") or "",
            "担当者":    person,
            "段取り":    _html_to_plain(t.get("plan") or "")[:200],
            "進捗":      _html_to_plain(t.get("recent") or "")[:400],
            "期限":      (t.get("deadline") or "").replace("-", "/"),
            "打ち上げ":  t.get("status_mark") or "",
            "工数(H)":   str(manhours) if manhours else "",
        })
    owner_label = name_filter or "(全員)"
    print(f"[INFO] 業務進捗表(Web): {len(tasks)} 行 (担当者={owner_label})", flush=True)
    return tasks


# ── AI 要約 ──────────────────────────────────────────────────────────
def _build_summary_prompt(records: list[dict], days: int,
                          nippo_text: str = "",
                          task_rows: list[dict] | None = None,
                          task_owner: str = "") -> str:
    """要約用プロンプトを構築（API 呼び出しにも、手貼りにも使用）

    記録を優先度別に3セクションに分けて配置:
      S1: 個人発信 (1:1 sent)   ← 最重要・先頭配置
      S2: グループ発信 (group/channel sent)
      S3: 被メンション + 1:1相手発言 (mentioned / other)
    """
    # 分類
    personal_sent = sorted(
        [r for r in records if r["subtype"] == "sent" and "1:1" in r["source"]],
        key=lambda x: x["dt"],
    )
    group_sent = sorted(
        [r for r in records if r["subtype"] == "sent" and "1:1" not in r["source"]],
        key=lambda x: x["dt"],
    )
    mentioned_other = sorted(
        [r for r in records if r["subtype"] in ("mentioned", "other")],
        key=lambda x: x["dt"],
    )

    def fmt(rec: dict, max_text: int = 180) -> str:
        return (
            f"[{rec['date']} {rec['time']}][{rec['source'][:35]}]"
            f" {rec['text'][:max_text]}"
        )

    sections = []

    if personal_sent:
        sections.append("### ★ 個人作業・先行開発（1:1チャット 自分発信）★")
        sections.append("# このセクションは最重要。自分が主体的に取り組んだ固有の技術作業が含まれる。")
        sections.extend(fmt(r, 200) for r in personal_sent)

    sections.append("### グループ・チャンネル 自分発信")
    sections.extend(fmt(r, 150) for r in group_sent[:220])

    sections.append("### 被メンション・1:1相手発言（参考文脈）")
    sections.extend(fmt(r, 80) for r in mentioned_other[:80])

    log_text = "\n".join(sections)

    # ─ 日報セクション ─
    nippo_section = ""
    if nippo_text.strip():
        nippo_section = f"\n\n---\n### 日報（業務実績抜粋）\n{nippo_text[:3000]}"

    # ─ 業務進捗表セクション ─
    task_section = ""
    task_update_instruction = ""
    if task_rows:
        owner_label = f"（担当：{task_owner}）" if task_owner else ""
        lines = [f"### 現在の業務進捗表{owner_label}"]
        lines.append("| # | 製品 | 業務ジャンル | 段取り(抜粋) | 進捗(現状) | 期限 | 打ち上げ | 工数(H) |")
        lines.append("|---|------|------|------|------|------|------|------|")
        for idx, t in enumerate(task_rows, 1):
            proc = t["進捗"][:120].replace("\n", " / ")
            plan = t["段取り"][:80].replace("\n", " / ")
            lines.append(
                f"| {idx} | {t['製品']} | {t['業務ジャンル']} | {plan} | {proc} "
                f"| {t['期限']} | {t['打ち上げ']} | {t['工数(H)']} |"
            )
        task_section = "\n\n---\n" + "\n".join(lines)
        task_update_instruction = """
## 業務進捗表 更新案
（上記「現在の業務進捗表」が提供された場合のみ出力すること）

### 既存行の更新案
今週のTeamsログ・日報を踏まえ、各行の「進捗」「期限」「工数(H)」の更新が必要な場合に提示する。
形式: **[#行番号 製品/業務ジャンル]** → F列進捗更新案・G列期限・H列工数の差分のみ記述

### 新規追加案
Teamsログ・日報に登場するが進捗表に未登録のタスクを以下の形式で提示:
| 製品 | 業務ジャンル | 担当者 | 段取り | 進捗 | 期限 | 打ち上げ | 工数(H) |"""

    return f"""以下は直近{days}日間のTeamsチャットログです。
{nippo_section}{task_section}

【セクションの意味】
- 「★ 個人作業・先行開発」: 1:1チャットで自分が発信した内容。自分が主体的に取り組んだ固有の技術作業。
- 「グループ・チャンネル 自分発信」: グループ/チャンネルで自分が送ったメッセージ（分析・報告・提案・依頼など）。
- 「被メンション・1:1相手発言」: 他者からの連絡・通知。自分が何かをした証拠ではない。

【主な取り組みトピックの選定基準 ─ 必ず守ること】
◎ 含める: 「★ 個人作業」または「自分発信」セクションに、以下の内容を含むメッセージがあるトピック
  - 技術的な調査・分析・検証の結果報告
  - 設計・仕様・コードの作成・提案・変更
  - 課題の特定・根本原因の分析
  - 自ら取り組んだ実験・評価の報告

✕ 除外: 「被メンション」のみで自分発信がない、または以下のみのトピック
  - 単純な受領返信（「了解」「確認しました」「ありがとうございます」）
  - 他者の作業進捗の報告受領（自分が何もしていない）
  - 会議の設定・日程調整・メンバー確認

→ つまり「自分は何をしたか？」が答えられないトピックは除外すること。

チャットログに登場する固有名詞・技術用語・人名・JIRAチケット番号はそのまま使用してください。
汎用的・抽象的な表現（「技術的な問題に取り組んだ」等）は避け、「自分が具体的に何をしたか」を記述してください。

## 主な取り組みトピック
（5〜8項目。「★個人作業」のテーマは必ず含める。被メンション受領のみのテーマは除外）

## 今週の成果・進捗
（3〜5項目。自分が完了させた・前進させたことのみ）

## 判明した課題・技術的知見
（3〜5項目。自分が発見・特定・分析した内容）

## 継続中の宿題・TODO
（3〜5項目。自分が対応中・次週やること）

## コミュニケーション・調整
（2〜4項目のみ。会議設定・日程調整・転送確認など。技術作業は上記に記載済み）
{task_update_instruction}
---
{log_text}
"""


def summarize_report(
    records: list[dict],
    days: int,
    api_key: str = "",
    endpoint: str = "",
    model: str = "",
    nippo_text: str = "",
    task_rows: list[dict] | None = None,
    task_owner: str = "",
) -> str | None:
    """
    LLM API でレコードを要約する。
    api_key が空なら OPENAI_API_KEY 環境変数を参照。
    どちらもなければ None を返す。

    endpoint を省略した場合:
      sk- 始まり → OpenAI API
      それ以外   → GitHub Models API

    model を省略した場合:
      GitHub Models → gpt-4o
      OpenAI        → gpt-4o-mini

    Note: gpt-4o-mini は GitHub Models 無料枠で入力 8K tokens 上限。
    プロンプトが大きい場合は 413 が返るため自動で gpt-4o にフォールバック。
    """
    import requests as _req

    key = _load_api_key(api_key)
    if not key:
        print("[SKIP] AI要約: API key なし (--api-key または tools/.gh_models_token を設定)", flush=True)
        return None

    if not endpoint:
        endpoint = (
            "https://api.openai.com/v1/chat/completions"
            if key.startswith("sk-")
            else "https://models.inference.ai.azure.com/chat/completions"
        )

    use_model = model.strip() or "gpt-4o"
    prompt = _build_summary_prompt(records, days,
                                   nippo_text=nippo_text,
                                   task_rows=task_rows,
                                   task_owner=task_owner)

    # プロキシ設定
    proxies = {}
    proxy_url = os.environ.get("HTTPS_PROXY") or os.environ.get("HTTP_PROXY")
    if proxy_url:
        proxies = {"http": proxy_url, "https": proxy_url}

    _FALLBACK_MODEL = "gpt-4o"
    for attempt_model in [use_model] + ([_FALLBACK_MODEL] if use_model != _FALLBACK_MODEL else []):
        print(f"  endpoint: {endpoint.split('/')[2]} / model: {attempt_model}", flush=True)
        try:
            resp = _req.post(
                endpoint,
                headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                json={"model": attempt_model,
                      "messages": [{"role": "user", "content": prompt}],
                      "max_tokens": 4000, "temperature": 0.3},
                proxies=proxies or None,
                timeout=120,
            )
            if resp.status_code == 200:
                return resp.json()["choices"][0]["message"]["content"].strip()
            # 413 = tokens_limit_reached → フォールバック対象
            err_code = ""
            try:
                err_code = resp.json().get("error", {}).get("code", "")
            except Exception:
                pass
            if resp.status_code == 413 or err_code == "tokens_limit_reached":
                print(f"  [INFO] {attempt_model} トークン上限超過 → フォールバック", flush=True)
                continue
            print(f"  [WARN] AI要約 {resp.status_code}: {resp.text[:300]}", flush=True)
            break
        except Exception as e:
            print(f"  [WARN] AI要約 例外: {e}", flush=True)
            break
    return None


# ── メイン ────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Teams チャット週報自動生成")
    ap.add_argument("--days",         type=int, default=7,    help="集計対象日数（デフォルト7）")
    ap.add_argument("--device-code",  action="store_true",    help="Device Code Flow 認証")
    ap.add_argument("--no-channels",  action="store_true",    help="Teams チャンネルをスキップ")
    ap.add_argument("--list-teams",   action="store_true",    help="参加中 Teams + チャンネル一覧を表示して終了")
    ap.add_argument("--teams-allow",  type=str, default="",   help="スキャン対象 Teams 名（部分一致・カンマ区切り）例: BEV,FCM,DCAP")
    ap.add_argument("--save-teams-allow", action="store_true", help="--teams-allow の値を tools/.teams_allow に保存")
    ap.add_argument("--out",         type=str, default="",    help="出力ファイルパス（省略時: 自動命名）")
    ap.add_argument("--summarize",   action="store_true",     help="AI要約を生成（API key 必要）")
    ap.add_argument("--prompt-only", action="store_true",     help="LLM に貼るプロンプトをファイル出力（API 不要）")
    ap.add_argument("--api-key",     type=str, default="",    help="LLM API key（省略時: tools/.gh_models_token または GH_MODELS_TOKEN 環境変数）")
    ap.add_argument("--save-key",    action="store_true",      help="--api-key で指定したキーを tools/.gh_models_token に保存して次回から省略可能にする")
    ap.add_argument("--endpoint",    type=str, default="",    help="OpenAI 互換エンドポイント URL（Azure OpenAI 等）")
    ap.add_argument("--model",       type=str, default="",    help="使用モデル名（例: gpt-4o-mini, gpt-4o, Llama-3.3-70B-Instruct）")
    ap.add_argument("--nippo-dir",   type=str, default="",    help="日報ディレクトリ（省略時: スクリプト/../日報/YYYY/）")
    ap.add_argument("--task-excel",  type=str, default="",    help="業務進捗Excelファイルのパス（省略時: tools/.user_config.jsonのtask_excel_path）")
    ap.add_argument("--task-source", choices=["excel", "web"], default="excel",
                     help="業務進捗データの取得元 (excel=Excelファイル直接読込 / web=manage_exp_progressのtasks.json)")
    ap.add_argument("--task-json",   type=str, default="",    help="manage_exp_progressのtasks.jsonのパス（--task-source web の時のみ使用。省略時: tools/.user_config.jsonのtask_json_path、それも無ければ manage_exp_progress/app/data/tasks.json）")
    ap.add_argument("--task-owner",  type=str, default="",    help="業務進捗表のフィルタ担当者名（省略時: tools/.user_config.jsonのtask_owner）")
    ap.add_argument("--save-task-owner", action="store_true", help="--task-owner で指定した値を tools/.user_config.json に保存して次回から省略可能にする")
    ap.add_argument("--from-cache",  action="store_true",
                     help="直前に収集したTeamsメッセージ(output/.records_cache.json)を再利用し、"
                          "Graph API再収集をスキップする（--summarize/--prompt-only を後から付けたい時に高速化）")
    args = ap.parse_args()

    _user_cfg = _load_user_config()
    if not args.task_owner:
        args.task_owner = _user_cfg.get("task_owner", "")
    if not args.task_excel:
        args.task_excel = _user_cfg.get("task_excel_path", "")
    if not args.task_json:
        args.task_json = _user_cfg.get("task_json_path", "")

    if args.save_key and args.api_key:
        _save_api_key(args.api_key)
    if args.save_task_owner and args.task_owner:
        _save_user_config(task_owner=args.task_owner)
    if args.save_teams_allow and args.teams_allow:
        _TEAMS_ALLOW_FILE.write_text(args.teams_allow.strip(), encoding="utf-8")
        print(f"[INFO] teams-allow を保存: {_TEAMS_ALLOW_FILE}", flush=True)

    auth_mode = "device_code" if args.device_code else "interactive"

    if args.from_cache:
        # ── キャッシュ再利用: Graph API認証・再収集をスキップ ──
        if not _RECORDS_CACHE_FILE.exists():
            print(f"[ERROR] キャッシュファイルが見つかりません: {_RECORDS_CACHE_FILE}\n"
                  f"        先に --from-cache 無しで一度実行してください。", flush=True)
            sys.exit(1)
        cache = _load_records_cache()
        records  = cache["records"]
        me_name  = cache["me_name"]
        since    = datetime.fromisoformat(cache["since"])
        args.days = cache["days"]  # 集計期間はキャッシュ側に合わせる
        print(f"[INFO] キャッシュを再利用します（{cache['saved_at']} 時点収集、"
              f"{cache['days']}日間、{len(records)}件）: {_RECORDS_CACHE_FILE}", flush=True)
    else:
        since = datetime.now(JST) - timedelta(days=args.days)
        print(f"[INFO] 集計期間: {since.strftime('%Y-%m-%d')} ～ 本日（JST）", flush=True)

        token = get_token(auth_mode)
        gc = GraphClient(token)

        # --list-teams: 一覧表示して終了
        if args.list_teams:
            list_teams(gc)
            sys.exit(0)

        # 自分のユーザー情報を取得
        me = gc.get("/me", params={"$select": "id,displayName,userPrincipalName"})
        if not me or not me.get("id"):
            print("[ERROR] /me API にアクセスできません", flush=True)
            sys.exit(1)
        me_id   = me["id"]
        me_name = me.get("displayName") or me.get("userPrincipalName", "")
        print(f"[INFO] ユーザー: {me_name} (id={me_id})", flush=True)

        # チャットメッセージ収集
        records = collect_chat_messages(gc, me_id, since)

        # チャンネルメッセージ収集
        teams_allow = _load_teams_allow(args.teams_allow)
        if not args.no_channels:
            records += collect_channel_messages(gc, me_id, since, teams_allow=teams_allow)

        print(f"\n[INFO] 収集完了: {len(records)} 件", flush=True)

        # 次回 --from-cache で再利用できるようキャッシュ保存
        _save_records_cache(records, args.days, me_name, me_id, since)

    # ─ 日報読み込み ─（週報本文・AI要約プロンプトの両方で使うため常に読み込む）
    nippo_text = ""
    _nippo_year = since.strftime("%Y")
    _nippo_base = (
        Path(args.nippo_dir) if args.nippo_dir
        else Path(__file__).parent.parent / "日報" / _nippo_year
    )
    if _nippo_base.exists():
        nippo_text = load_nippo_text(_nippo_base, since, datetime.now(JST))
        print(f"[INFO] 日報: {len(nippo_text)} 文字 読み込み完了", flush=True)
    else:
        print(f"[INFO] 日報ディレクトリが見つかりません: {_nippo_base}", flush=True)

    # ─ 業務進捗表読み込み (Excel 直接 or manage_exp_progressのtasks.json) ─
    task_rows: list[dict] | None = None
    if args.summarize or args.prompt_only:
        if args.task_source == "web":
            json_path = args.task_json or str(
                Path(__file__).parent.parent / "manage_exp_progress" / "app" / "data" / "tasks.json")
            task_rows = load_task_json(json_path, name_filter=args.task_owner)
        elif args.task_excel:
            task_rows = load_task_excel(args.task_excel, name_filter=args.task_owner)

    # ─ AI 要約 or プロンプト出力 ─
    ai_summary = None
    if args.summarize:
        print("\n[INFO] AI要約を生成中...", flush=True)
        ai_summary = summarize_report(
            records, args.days,
            api_key=args.api_key,
            endpoint=args.endpoint,
            model=args.model,
            nippo_text=nippo_text,
            task_rows=task_rows,
            task_owner=args.task_owner,
        )

    if args.prompt_only or args.summarize:
        # --summarize 時も常に保存（AI入力の監査・デバッグ用）
        stamp = datetime.now(JST).strftime("%Y%m%d_%H%M")
        out_dir2 = Path(__file__).parent.parent / "output"
        out_dir2.mkdir(exist_ok=True)
        prompt_path = out_dir2 / f"summary_prompt_{stamp}.txt"
        prompt_path.write_text(
            _build_summary_prompt(records, args.days,
                                  nippo_text=nippo_text,
                                  task_rows=task_rows,
                                  task_owner=args.task_owner),
            encoding="utf-8")
        if args.prompt_only or ai_summary is None:
            print(f"\n[INFO] 要約プロンプトを出力しました (ChatGPT/Claude に貼ってください):\n  {prompt_path.resolve()}", flush=True)
        else:
            print(f"\n[INFO] 要約プロンプトを保存しました: {prompt_path.name}", flush=True)

    # 週報生成（AI要約がある場合はログ羅列をスキップ）
    report = build_report(records, me_name, args.days, ai_summary=ai_summary or "", nippo_text=nippo_text)

    # 出力先決定
    out_dir = Path(__file__).parent.parent / "output"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now(JST).strftime("%Y%m%d_%H%M")
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = out_dir / f"weekly_report_{stamp}.md"

    out_path.write_text(report, encoding="utf-8")
    print(f"\n[DONE] 週報候補案を出力しました:\n  {out_path.resolve()}", flush=True)

    # 業務進捗表 更新案 を専用ファイルにも保存
    _marker = "## 業務進捗表 更新案"
    if ai_summary and _marker in ai_summary:
        _task_text = ai_summary[ai_summary.index(_marker):]
        _task_path = out_dir / f"task_update_{stamp}.md"
        _task_path.write_text(_task_text, encoding="utf-8")
        print(f"[DONE] 業務進捗表 更新案を保存しました:\n  {_task_path.resolve()}", flush=True)

    # コンソールにも表示
    print("\n" + "=" * 60, flush=True)
    print(report[:3000], flush=True)
    if len(report) > 3000:
        print(f"... (以下省略 — ファイルを参照してください)", flush=True)

if __name__ == "__main__":
    main()
